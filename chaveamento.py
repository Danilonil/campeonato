
import PySimpleGUI as sg
sg.theme('SystemDefaultForReal')

from random import shuffle
from time import sleep
from openpyxl import load_workbook, Workbook
lista = load_workbook('Lista Geral.xlsx')
competidores = lista['competidores']

def tela_inicial():
    Layout = [
    [sg.Text('ID:    ', size = (6,1)), sg.InputText(key = 'id_aluno',size = (14,1), do_not_clear=False), 
    sg.Text('Categoria:'), sg.InputText(key = 'categ', size = (22,1), do_not_clear=False)],
    [sg.Text('Nome:  '), sg.InputText(key = 'nome', size = (49,1), do_not_clear=False)],
    [sg.Text('Escola:'), sg.InputText(key = 'escola', size = (49,1), do_not_clear=False)],
    [sg.Text(key = 'validação')],
    [sg.Button('Remover', size = (20), pad = (30,0)), sg.Button('Adicionar', size = (20), pad = (0,0))],
    [sg.Text('_'*58)],
    [sg.Text('Últimos adicionados', size = (49,1))],
    [sg.Combo(['Pré-mirim Masc.', 'Pré-mirim Fem.','Mirim Masc.', 'Mirim Fem.', 'Infantil Masc.', 'Infantil Fem.'], enable_events=True, key='combo'), sg.Button('Ver lista')],
    [sg.Text('_'*58)],
    [sg.Text('Selecione a chave:', pad = (40,0)), sg.Radio('Chave - A', group_id = 'radio1', key = 'A'), 
    sg.Radio('Chave - B', group_id = 'radio1', key = 'B')],
    [sg.Text(key = 'validação2')],
    [sg.Button('Iniciar chaveamento', size = (30,1), pad = (40,1), button_color= 'LIGHT BLUE4', font = ("Arial", 14))],
    [sg.Text(key = 'validação3')]
    ]

    return sg.Window('Chaveamento ultra blaster', Layout, finalize = True)

def popup():
    layout7 = [[sg.Text('O chaveamento deve ser iniciado somente depois que todos os competidores forem adicionados ao torneio. \n Esta ação será permanente. Deseja continuar?')], 
    [sg.Button('NÃO', size = (20), pad = (100,0) , button_color = 'LIGHT BLUE4'), 
    sg.Button('SIM', size = (20), button_color = 'LIGHT BLUE4')]]
    return sg.Window('ATENÇÃO', layout7, finalize = True)

janela = tela_inicial()
janela2 = None

janela['id_aluno'].bind("<Return>", "_Enter")
pmm = []
pmf = []
mm = []
mf = []
im = []
inf = []
listaMae = [pmm, pmf, mm, mf, im, inf]
num = 0
while True:
    window, evento, valor = sg.read_all_windows()
    if window == janela and evento == sg.WIN_CLOSED:
        break

    for linha in competidores.iter_rows(min_row = 2, values_only = True):        
        if evento == "id_aluno" + "_Enter":
            if str(linha[4]) in valor['id_aluno']:
                janela['categ'].Update(f'{linha[0]} {linha[1]}')
                janela['nome'].Update(f'{linha[5].title()}')
                janela['escola'].Update(f'{linha[3].title()}')
                num = valor['id_aluno']
                janela['validação'].Update('')
                break
            else:
                janela['validação'].Update('ID INVALIDO'.center(93), text_color = 'RED')
                num = valor['id_aluno']

        if evento == 'Adicionar' and str(linha[4]) in str(num):
            if linha[0] == 'Pré-Mirim' and linha[1] == 'Masculino':
                if len(pmm) <1:
                    pmm.append([linha[4], linha[5], linha[3]])
                    janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                    num = 0            
                else:
                    for n in pmm:
                        if num in str(n[0]):
                            janela['validação'].Update('ALUNO JÁ FOI CADASTRADO ANTES'.center(72), text_color = 'RED')
                            break
                    else:
                        pmm.append([linha[4], linha[5], linha[3]])
                        janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                        num = 0
                        break

            elif linha[0] == 'Pré-Mirim' and linha[1] == 'Feminino':
                if len(pmf) <1:
                    pmf.append([linha[4], linha[5], linha[3]])
                    janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                    num = 0            
                else:
                    for n in pmf:
                        if num in str(n[0]):
                            janela['validação'].Update('ALUNA JÁ FOI CADASTRADA ANTES'.center(72), text_color = 'RED')
                            break
                    else:
                        pmf.append([linha[4], linha[5], linha[3]])
                        janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                        num = 0
                        break

            elif linha[0] == 'Mirim' and linha[1] == 'Masculino':
                if len(mm) <1:
                    mm.append([linha[4], linha[5], linha[3]])
                    janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                    num = 0            
                else:
                    for n in mm:
                        if num in str(n[0]):
                            janela['validação'].Update('ALUNO JÁ FOI CADASTRADO ANTES'.center(72), text_color = 'RED')
                            break
                    else:
                        mm.append([linha[4], linha[5], linha[3]])
                        janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                        num = 0
                        break

            elif linha[0] == 'Mirim' and linha[1] == 'Feminino':
                if len(mf) <1:
                    mf.append([linha[4], linha[5], linha[3]])
                    janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                    num = 0            
                else:
                    for n in mf:
                        if num in str(n[0]):
                            janela['validação'].Update('ALUNA JÁ FOI CADASTRADA ANTES'.center(72), text_color = 'RED')
                            break
                    else:
                        mf.append([linha[4], linha[5], linha[3]])
                        janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                        num = 0
                        break

            elif linha[0] == 'Infantil' and linha[1] == 'Masculino':
                if len(im) <1:
                    im.append([linha[4], linha[5], linha[3]])
                    janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                    num = 0            
                else:
                    for n in im:
                        if num in str(n[0]):
                            janela['validação'].Update('ALUNO JÁ FOI CADASTRADO ANTES'.center(72), text_color = 'RED')
                            break
                    else:
                        im.append([linha[4], linha[5], linha[3]])
                        janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                        num = 0
                        break

            elif linha[0] == 'Infantil' and linha[1] == 'Feminino':
                if len(inf) <1:
                    inf.append([linha[4], linha[5], linha[3]])
                    janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                    num = 0            
                else:
                    for n in inf:
                        if num in str(n[0]):
                            janela['validação'].Update('ALUNA JÁ FOI CADASTRADA ANTES'.center(72), text_color = 'RED')
                            break
                    else:
                        inf.append([linha[4], linha[5], linha[3]])
                        janela['validação'].Update('Adicionado com sucesso'.center(86), text_color = 'GREEN')
                        num = 0
                        break

        if evento == 'Remover' and str(linha[4]) == num:
            for pos, item in enumerate(listaMae):
                for p, i in enumerate(item):
                    if num in str(i[0]):
                        del listaMae[pos][p]
                        janela['validação'].Update('Deletado com sucesso'.center(86), text_color = 'GREEN')
            else:
                num = 0 
    
    if evento == 'Ver lista':
        combo = valor['combo']
        if combo == 'Pré-mirim Masc.':
            layout1 = []
            titulo = [[sg.Text('ID', size=(8,1)), sg.Text('NOME', size=(40)), sg.Text('ESCOLA', size=(40))]]
            for i in pmm:
                lista = [[sg.Text(i[0], size = (8)), sg.Text(i[1].title(), size = (40)), sg.Text(i[2].title(), size = (40))]]
                layout1 = layout1 + lista
            layout1 = [[sg.Column(titulo)],
            [sg.Column(layout1, scrollable= True, vertical_scroll_only= True, )]]
            window = sg.Window('Tabela de competidores', layout1)
            event, values = window.read()

        elif combo == 'Pré-mirim Fem.':
            layout2 = []
            titulo = [[sg.Text('ID', size=(8,1)), sg.Text('NOME', size=(40)), sg.Text('ESCOLA', size=(40))]]
            for i in pmf:
                lista = [[sg.Text(i[0], size = (8)), sg.Text(i[1].title(), size = (40)), sg.Text(i[2].title(), size = (40))]]
                layout2 = layout2 + lista
            layout2 = [[sg.Column(titulo)],
            [sg.Column(layout2, scrollable= True, vertical_scroll_only= True, )]]
            window = sg.Window('Tabela de competidores', layout2)
            event, values = window.read()

        elif combo == 'Mirim Masc.':
            layout3 = []
            titulo = [[sg.Text('ID', size=(8,1)), sg.Text('NOME', size=(40)), sg.Text('ESCOLA', size=(40))]]
            for i in mm:
                lista = [[sg.Text(i[0], size = (8)), sg.Text(i[1].title(), size = (40)), sg.Text(i[2].title(), size = (40))]]
                layout3 = layout3 + lista 
            layout3 = [[sg.Column(titulo)],
            [sg.Column(layout3, scrollable= True, vertical_scroll_only= True, )]]
            window = sg.Window('Tabela de competidores', layout3)
            event, values = window.read()

        elif combo == 'Mirim Fem.':
            layout4 = []
            titulo = [[sg.Text('ID', size=(8,1)), sg.Text('NOME', size=(40)), sg.Text('ESCOLA', size=(40))]]
            for i in mf:
                lista = [[sg.Text(i[0], size = (8)), sg.Text(i[1].title(), size = (40)), sg.Text(i[2].title(), size = (40))]]
                layout4 = layout4 + lista
            layout4 = [[sg.Column(titulo)],
            [sg.Column(layout4, scrollable= True, vertical_scroll_only= True, )]]
            window = sg.Window('Tabela de competidores', layout4)
            event, values = window.read()

        elif combo == 'Infantil Masc.':
            layout5 = []
            titulo = [[sg.Text('ID', size=(8,1)), sg.Text('NOME', size=(40)), sg.Text('ESCOLA', size=(40))]]
            for i in im:
                lista = [[sg.Text(i[0], size = (8)), sg.Text(i[1].title(), size = (40)), sg.Text(i[2].title(), size = (40))]]
                layout5 = layout5 + lista
            layout5 = [[sg.Column(titulo)],
            [sg.Column(layout5, scrollable= True, vertical_scroll_only= True, )]] 
            window = sg.Window('Tabela de competidores', layout5)
            event, values = window.read()

        elif combo == 'Infantil Fem.':
            layout6 = []
            titulo = [[sg.Text('ID', size=(8,1)), sg.Text('NOME', size=(40)), sg.Text('ESCOLA', size=(40))]]
            for i in inf:
                lista = [[sg.Text(i[0], size = (8)), sg.Text(i[1].title(), size = (40)), sg.Text(i[2].title(), size = (40))]]
                layout6 = layout6 + lista 
            layout6 = [[sg.Column(titulo)],
            [sg.Column(layout6, scrollable= True, vertical_scroll_only= True, )]]
            window = sg.Window('Tabela de competidores', layout6)
            event, values = window.read()
    
    if window == janela:
        if valor['A'] == True or valor['B'] == True:
            chave = 'B'
            if valor['A'] == True:
                chave = 'A'
            if evento == 'Iniciar chaveamento':
                janela['validação2'].Update('')
                janela['validação'].Update('')
                janela2 = popup()
        elif evento == 'Iniciar chaveamento':
            janela['validação2'].Update('Escolha uma CHAVE antes de prosseguir'.center(77), text_color = 'RED')
            janela['validação'].Update('')


    if window == janela2 and evento == sg.WIN_CLOSED or evento == 'NÃO':
        janela2.hide()

    if  evento == 'SIM':
        janela2.hide()
        planilha = Workbook()
        categoria = ['Pré-Mirim Masc.', 'Pré-Mirim Fem.', 'Mirim Masc.', 'Mirim Fem.', 'Infantil Masc.', 'Infantil Fem.']
        del planilha['Sheet']
        for c in categoria:
            planilha.create_sheet(c)
            planilha[c].cell(row = 1, column = 1, value = c)
        for pagina in planilha:
            pagina.column_dimensions['B'].width = 50
            pagina.column_dimensions['C'].width = 50
            pagina.merge_cells('A1:C1')
            pagina.append(['ID', 'NOME', 'ESCOLA'])
        for item in listaMae:
            for jogador in item:
                if item == pmm:
                    planilha['Pré-Mirim Masc.'].append([jogador[0], jogador[1], jogador[2]])
                elif item == pmf:
                    planilha['Pré-Mirim Fem.'].append([jogador[0], jogador[1], jogador[2]])
                elif item == mm:
                    planilha['Mirim Masc.'].append([jogador[0], jogador[1], jogador[2]])
                elif item == mf:
                    planilha['Mirim Fem.'].append([jogador[0], jogador[1], jogador[2]])
                elif item == im:
                    planilha['Infantil Masc.'].append([jogador[0], jogador[1], jogador[2]])
                elif item == inf:
                    planilha['Infantil Fem.'].append([jogador[0], jogador[1], jogador[2]])

        planilha.save('Competidores.xlsx')
        dados = []
        padrão = []
        potencia = 1
        for item in listaMae:
            shuffle(item)
            dados.append(len(item))

        while True:
            n = 2 ** potencia
            padrão.append(n)
            potencia += 1
            if max(dados) <= n:
                break

        barra = sum(dados)
        progresso = 0
        for i in range(barra+1):
            if barra == 0:
                break
            sg.one_line_progress_meter(title = '', current_value = i + progresso, max_value = barra, 
            orientation = 'h', no_button = True, no_titlebar = True)
            sleep(0.01)

        wb = load_workbook('sumula.xlsx')
        for item in listaMae:
            contjogador = isentos = 0
            if item == pmm:
                planilha1 = wb['Pré-Mirim Masc.']
                for linha in range(2, len(pmm)//8 + 4):
                    planilha1.cell(row = linha, column = 2, value = 'PRÉ-MIRIM')
                    planilha1.cell(row = linha, column = 3, value = 'MASCULINO')
            elif item == pmf:
                planilha1 = wb['Pré-Mirim Fem.']
                for linha in range(2, len(pmf)//8 + 4):
                    planilha1.cell(row = linha, column = 2, value = 'PRÉ-MIRIM')
                    planilha1.cell(row = linha, column = 3, value = 'FEMININO')
            elif item == mm:
                planilha1 = wb['Mirim Masc.']
                for linha in range(2, len(mm)//8 + 4):
                    planilha1.cell(row = linha, column = 2, value = 'MIRIM')
                    planilha1.cell(row = linha, column = 3, value = 'MASCULINO')
            elif item == mf:
                planilha1 = wb['Mirim Fem.']
                for linha in range(2, len(mf)//8 + 4):
                    planilha1.cell(row = linha, column = 2, value = 'MIRIM')
                    planilha1.cell(row = linha, column = 3, value = 'FEMININO')
            elif item == im:
                planilha1 = wb['Infantil Masc.']
                for linha in range(2, len(im)//8 + 4):
                    planilha1.cell(row = linha, column = 2, value = 'INFANTIL')
                    planilha1.cell(row = linha, column = 3, value = 'MASCULINO')
            elif item == inf:
                planilha1 = wb['Infantil Fem.']
                for linha in range(2, len(inf)//8 + 4):
                    planilha1.cell(row = linha, column = 2, value = 'INFANTIL')
                    planilha1.cell(row = linha, column = 3, value = 'FEMININO')

            lin = 2
            col = 4
            for jogador in item:
                if col == 20:
                    lin += 1
                    col = 4
                if planilha1.cell(row = lin, column = 1, value = ''):
                    planilha1.cell(row = lin, column = 1, value = chave)
                planilha1.cell(row = lin, column = col, value = jogador[1])
                col += 1
                planilha1.cell(row = lin, column = col, value = jogador[2])
                col += 1 
                contjogador += 1
                progresso += 1

                for p in padrão:
                    if p < len(item) < p *2:
                        if len(item) - p > (p *2) - len(item):
                            isentos = (p *2) - len(item)
                        else:
                            isentos = len(item) - p

                if contjogador == len(item) - isentos :
                    break
            
            for jogador in item[contjogador:]:
                if col == 20:
                    lin += 1
                    col = 4
                if planilha1.cell(row = lin, column = 1, value = ''):
                    planilha1.cell(row = lin, column = 1, value = chave)
                planilha1.cell(row = lin, column = col, value = jogador[1])
                col += 1
                planilha1.cell(row = lin, column = col, value = jogador[2])
                col += 3
                progresso += 1

        wb.save('Mala Direta - MM.xlsx')
        janela['validação3'].Update('Chaveamento finalizado'.center(82), font = ("Arial", 12), text_color = 'GREEN')

janela.Close()
